"""
Tarea 7.1 — Servicio de generación de reportes.

Soporta formatos: PDF (fpdf2), Excel (openpyxl), CSV (csv stdlib), JSON.
Los reportes se generan de forma asíncrona a través de Celery y el resultado
se sube a S3. Este módulo implementa la lógica de construcción del reporte
a partir de los datos consultados en MongoDB.
"""

import csv
import io
import json
import logging
from datetime import datetime
from typing import Any

logger = logging.getLogger(__name__)


# ─────────────────────────────────────────────────────────────────────────────
# Consultores de datos por tipo de reporte
# ─────────────────────────────────────────────────────────────────────────────


async def _fetch_activity_data(
    db, filters: dict, date_from: datetime | None, date_to: datetime | None
) -> list[dict]:
    query: dict = {}
    if date_from:
        query.setdefault("timestamp", {})["$gte"] = date_from
    if date_to:
        query.setdefault("timestamp", {})["$lte"] = date_to
    if filters.get("user_id"):
        query["user_id"] = filters["user_id"]

    rows = []
    async for log in db["audit_logs"].find(query).sort("timestamp", -1).limit(5000):
        rows.append(
            {
                "timestamp": str(log.get("timestamp", "")),
                "user_id": str(log.get("user_id", "")),
                "action": log.get("action", ""),
                "entity": log.get("entity_type", ""),
                "ip": log.get("ip_address", ""),
            }
        )
    return rows


async def _fetch_documents_data(
    db, filters: dict, date_from: datetime | None, date_to: datetime | None
) -> list[dict]:
    query: dict = {"is_active": True}
    if date_from:
        query.setdefault("created_at", {})["$gte"] = date_from
    if date_to:
        query.setdefault("created_at", {})["$lte"] = date_to

    rows = []
    async for doc in db["documents"].find(query).sort("created_at", -1).limit(5000):
        rows.append(
            {
                "id": str(doc.get("_id", "")),
                "title": doc.get("title", ""),
                "filename": doc.get("filename", ""),
                "owner_id": doc.get("owner_id", ""),
                "size_bytes": doc.get("size_bytes", 0),
                "created_at": str(doc.get("created_at", "")),
                "tags": ", ".join(doc.get("tags", [])),
            }
        )
    return rows


async def _fetch_workflows_data(
    db, filters: dict, date_from: datetime | None, date_to: datetime | None
) -> list[dict]:
    query: dict = {}
    if filters.get("status"):
        query["status"] = filters["status"]

    rows = []
    async for inst in db["workflow_instances"].find(query).limit(5000):
        rows.append(
            {
                "id": str(inst.get("_id", "")),
                "policy_id": inst.get("policy_id", ""),
                "status": inst.get("status", ""),
                "started_by": inst.get("started_by", ""),
                "current_node": inst.get("current_node_id", ""),
                "steps_completed": len(inst.get("history", [])),
            }
        )
    return rows


_DATA_FETCHERS = {
    "activity": _fetch_activity_data,
    "documents": _fetch_documents_data,
    "workflows": _fetch_workflows_data,
}


# ─────────────────────────────────────────────────────────────────────────────
# Generadores de formato
# ─────────────────────────────────────────────────────────────────────────────


def _to_csv(rows: list[dict]) -> bytes:
    if not rows:
        return b"No data"
    buf = io.StringIO()
    writer = csv.DictWriter(buf, fieldnames=list(rows[0].keys()))
    writer.writeheader()
    writer.writerows(rows)
    return buf.getvalue().encode("utf-8")


def _to_json(rows: list[dict]) -> bytes:
    return json.dumps(rows, ensure_ascii=False, indent=2, default=str).encode("utf-8")


def _to_excel(rows: list[dict], title: str) -> bytes:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = title[:30]

    if not rows:
        ws.append(["No data"])
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    headers = list(rows[0].keys())
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header.upper())
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for row_idx, row in enumerate(rows, 2):
        for col_idx, key in enumerate(headers, 1):
            ws.cell(row=row_idx, column=col_idx, value=str(row.get(key, "")))

    # Auto-width
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 60)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _to_pdf(rows: list[dict], title: str) -> bytes:
    """
    Genera un PDF básico con fpdf2.
    Si fpdf2 no está instalado, devuelve un CSV plano como fallback.
    """
    try:
        from fpdf import FPDF

        pdf = FPDF()
        pdf.add_page()
        pdf.set_font("Helvetica", "B", 16)
        pdf.cell(0, 10, title, ln=True, align="C")
        pdf.set_font("Helvetica", "", 8)
        pdf.ln(4)

        if not rows:
            pdf.cell(0, 10, "No data available.", ln=True)
        else:
            headers = list(rows[0].keys())
            col_w = max(10, 190 // len(headers))
            # Header row
            pdf.set_fill_color(31, 78, 121)
            pdf.set_text_color(255, 255, 255)
            pdf.set_font("Helvetica", "B", 7)
            for h in headers:
                pdf.cell(col_w, 7, h.upper(), border=1, fill=True)
            pdf.ln()
            # Data rows
            pdf.set_text_color(0, 0, 0)
            pdf.set_font("Helvetica", "", 7)
            for row in rows[:500]:  # límite de filas para PDF
                for key in headers:
                    pdf.cell(col_w, 6, str(row.get(key, ""))[:30], border=1)
                pdf.ln()

        return pdf.output()

    except ImportError:
        logger.warning("fpdf2 not installed, falling back to CSV for PDF report")
        return _to_csv(rows)


# ─────────────────────────────────────────────────────────────────────────────
# Servicio principal
# ─────────────────────────────────────────────────────────────────────────────


class ReportGenerator:
    """Genera reportes en múltiples formatos a partir de consultas MongoDB."""

    CONTENT_TYPES = {
        "pdf": "application/pdf",
        "excel": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "csv": "text/csv",
        "json": "application/json",
    }

    async def generate(
        self,
        report_type: str,
        fmt: str,
        title: str,
        filters: dict[str, Any] | None = None,
        date_from: datetime | None = None,
        date_to: datetime | None = None,
    ) -> tuple[bytes, str]:
        """
        Devuelve (content_bytes, content_type).
        """
        from app.core.database import get_database

        db = get_database()
        filters = filters or {}

        fetcher = _DATA_FETCHERS.get(report_type)
        if fetcher is None:
            rows: list[dict] = []
            logger.warning("No fetcher for report_type=%s, returning empty", report_type)
        else:
            rows = await fetcher(db, filters, date_from, date_to)

        if fmt == "pdf":
            content = _to_pdf(rows, title)
        elif fmt == "excel":
            content = _to_excel(rows, title)
        elif fmt == "csv":
            content = _to_csv(rows)
        else:
            content = _to_json(rows)

        content_type = self.CONTENT_TYPES.get(fmt, "application/octet-stream")
        logger.info("Report generated: type=%s, fmt=%s, rows=%d", report_type, fmt, len(rows))
        return content, content_type
