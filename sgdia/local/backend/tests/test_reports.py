"""
tests/test_reports.py — Tests para el generador de reportes.
"""

import pytest
from app.services.report_generator import _to_csv, _to_excel, _to_json

# ─────────────────────────────────────────────────────────────────────────────
# Unit tests del generador (sin DB)
# ─────────────────────────────────────────────────────────────────────────────

SAMPLE_ROWS = [
    {"id": "1", "user_id": "u1", "action": "login", "timestamp": "2026-01-01"},
    {"id": "2", "user_id": "u2", "action": "upload", "timestamp": "2026-01-02"},
    {"id": "3", "user_id": "u1", "action": "logout", "timestamp": "2026-01-03"},
]


def test_csv_output_has_headers():
    """CSV debe contener las cabeceras como primera línea."""
    content = _to_csv(SAMPLE_ROWS)
    text = content.decode("utf-8")
    first_line = text.split("\n")[0]
    assert "id" in first_line
    assert "action" in first_line


def test_csv_output_has_correct_row_count():
    """CSV debe tener 1 fila de cabecera + N filas de datos."""
    content = _to_csv(SAMPLE_ROWS)
    lines = [line for line in content.decode("utf-8").split("\n") if line.strip()]
    assert len(lines) == len(SAMPLE_ROWS) + 1


def test_csv_empty_rows():
    """CSV con lista vacía devuelve 'No data'."""
    content = _to_csv([])
    assert content == b"No data"


def test_json_output_valid():
    """JSON debe ser parseable y contener todos los registros."""
    import json

    content = _to_json(SAMPLE_ROWS)
    parsed = json.loads(content.decode("utf-8"))
    assert len(parsed) == 3
    assert parsed[0]["action"] == "login"


def test_excel_output_is_bytes():
    """Excel debe devolver bytes (archivo .xlsx)."""
    content = _to_excel(SAMPLE_ROWS, "Test Report")
    assert isinstance(content, bytes)
    assert len(content) > 0


def test_excel_can_be_opened():
    """El Excel generado debe poder abrirse con openpyxl."""
    import io

    import openpyxl

    content = _to_excel(SAMPLE_ROWS, "Activity Report")
    wb = openpyxl.load_workbook(io.BytesIO(content))
    ws = wb.active

    # Primera fila = headers
    headers = [cell.value for cell in ws[1]]
    assert "ID" in headers or "id".upper() in [h.upper() for h in headers if h]

    # Debe haber datos
    assert ws.max_row > 1


@pytest.mark.asyncio
async def test_report_generator_activity(mock_database):
    """ReportGenerator.generate() con type='activity' devuelve bytes."""
    from app.services.report_generator import ReportGenerator

    async def empty_cursor(*a, **kw):
        return
        yield

    cursor = MagicMock(__aiter__=empty_cursor)
    cursor.sort.return_value = cursor
    cursor.limit.return_value = cursor
    mock_database["audit_logs"].find.return_value = cursor

    gen = ReportGenerator()
    content, ctype = await gen.generate(
        report_type="activity",
        fmt="csv",
        title="Test Activity Report",
    )

    assert isinstance(content, bytes)
    assert "csv" in ctype or "text" in ctype


from unittest.mock import MagicMock
